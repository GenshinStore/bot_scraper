import asyncio
import json
import os
import sys
import random
import traceback
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl import Workbook

from dotenv import load_dotenv
from telethon import TelegramClient, events
from telethon.errors import UserPrivacyRestrictedError, UserNotMutualContactError, FloodWaitError, RPCError, SessionPasswordNeededError, UsersTooMuchError
from telethon.tl.functions.channels import InviteToChannelRequest
from telethon.tl.types import Channel, Chat, User, DocumentAttributeFilename

# Muat environment variables dari file .env
load_dotenv()

# =====================================================
# KONFIGURASI PATH & DASAR
# =====================================================
# Dapatkan direktori tempat skrip ini berada untuk path yang andal
SCRIPT_DIR = Path(__file__).resolve().parent

API_ID = int(os.getenv("API_ID", 0))
API_HASH = os.getenv("API_HASH")
BOT_TOKEN = os.getenv("BOT_TOKEN")
ADMIN_ID_STR = os.getenv("ADMIN_ID")
# Ganti dengan ID Telegram numerik Anda untuk mengamankan perintah admin
# Anda bisa mendapatkan ID Anda dari bot seperti @userinfobot
ADMIN_ID = int(ADMIN_ID_STR) if ADMIN_ID_STR and ADMIN_ID_STR.isdigit() else 0

SESSIONS_DIR = SCRIPT_DIR / "sessions"
BOT_SESSION_NAME = "bot_session" # Hanya nama, bukan path
BOT_SESSION_PATH = str(SESSIONS_DIR / BOT_SESSION_NAME)
HISTORY_DIR = SCRIPT_DIR / "history"
DEFAULT_HISTORY_FILE = HISTORY_DIR / "global_broadcast_history.xlsx"

os.makedirs(HISTORY_DIR, exist_ok=True)

# Pastikan direktori sesi ada
os.makedirs(SESSIONS_DIR, exist_ok=True)

# Hapus session lama jika ada untuk force fresh login
bot_session_file = f"{BOT_SESSION_PATH}.session"
if os.path.exists(bot_session_file):
    try:
        os.remove(bot_session_file)
        print("[INFO] Old bot session cleared for fresh login.")
    except Exception as e:
        print(f"[WARNING] Tidak bisa menghapus session lama: {e}")

# Klien bot utama
bot_client = TelegramClient(BOT_SESSION_PATH, API_ID, API_HASH)

# =====================================================
# STATUS GLOBAL
# =====================================================
# Dictionary untuk mengelola tugas yang berjalan per sesi
# Key: session_name, Value: { "running": bool, "task_name": str, "stop_requested": bool }
TASK_STATE = {}

# =====================================================
# FUNGSI HELPER RIWAYAT (HISTORY)
# =====================================================
def load_history_data(history_path: Path):
    """
    Memuat file riwayat dan mengembalikan peta status dan satu set semua ID yang diproses.
    Returns:
        tuple: (status_map, processed_ids)
        status_map (dict): {user_id: status}
        processed_ids (set): {user_id1, user_id2, ...}
    """
    if not history_path.exists():
        return {}, set()

    processed_ids = set()
    status_map = {}
    try:
        wb = openpyxl.load_workbook(history_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        try:
            uid_col_idx = headers.index("User ID")
            status_col_idx = headers.index("Status")
        except ValueError:
            print(f"[WARNING] File riwayat '{history_path.name}' memiliki header tidak valid. Melewatkan.")
            return {}, set()

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[uid_col_idx]: continue
            try:
                uid = int(row[uid_col_idx])
                status = row[status_col_idx]
                processed_ids.add(uid)
                status_map[uid] = status
            except (ValueError, TypeError, IndexError):
                continue
        return status_map, processed_ids
    except Exception as e:
        print(f"[ERROR] Gagal membaca file riwayat {history_path}: {e}")
        return {}, set()

def append_logs_to_history(history_path: Path, history_logs: list):
    """Menambahkan daftar entri log ke file riwayat Excel yang ditentukan."""
    if not history_logs: return

    headers = ["Timestamp", "User ID", "User Name", "Username", "Target Group ID", "Target Group Title", "Status", "Details"]
    
    try:
        wb = openpyxl.load_workbook(history_path) if history_path.exists() else Workbook()
        ws = wb.active
        if not history_path.exists() or ws.max_row == 0:
            ws.title = "Global Broadcast History"
            ws.append(headers)
        for log_entry in history_logs:
            ws.append([log_entry.get(h.lower().replace(" ", "_")) for h in headers])
        wb.save(history_path)
        print(f"[INFO] Menambahkan {len(history_logs)} catatan ke file riwayat: {history_path.name}")
    except Exception as e:
        print(f"[ERROR] Gagal menulis ke file riwayat {history_path}: {e}")

# =====================================================
# FUNGSI HELPER & LOGIKA INTI
# =====================================================

async def add_user_to_group(user_client, target_group, user_id, user_username=None):
    """Menambahkan user langsung ke grup, menangani supergroup dan grup dasar."""
    try:
        # PERBAIKAN: Resolve entitas user terlebih dahulu.
        # Ini penting agar client yang sedang berjalan "mengenal" user target,
        # terutama jika akun yang menambahkan berbeda dengan akun yang melakukan scrape.
        user_to_add = None
        try:
            # Cara tercepat adalah via ID jika sudah ada di cache sesi.
            user_to_add = await user_client.get_entity(user_id)
        except (ValueError, TypeError):
            # Jika gagal, coba cari via username jika tersedia.
            if user_username and user_username != 'N/A':
                user_to_add = await user_client.get_entity(user_username)
        
        # Jika user masih tidak ditemukan, biarkan request di bawah gagal agar errornya tercatat.
        if not user_to_add:
            user_to_add = user_id

        # Cek tipe grup untuk menggunakan request yang benar
        if isinstance(target_group, Channel): # Ini adalah Supergroup
            await user_client(InviteToChannelRequest(channel=target_group, users=[user_to_add]))
            return True, "Undangan terkirim" # Di supergrup, ini adalah undangan, bukan penambahan paksa.
        elif isinstance(target_group, Chat): # Ini adalah Grup Dasar
            from telethon.tl.functions.messages import AddChatUserRequest
            await user_client(AddChatUserRequest(chat_id=target_group.id, user_id=user_to_add, fwd_limit=10))
            return True, "Berhasil ditambahkan" # Di grup dasar, ini biasanya penambahan langsung.
        else:
            return False, "Tipe grup tidak didukung"

    except FloodWaitError as e:
        return False, f"FLOOD_WAIT:{e.seconds}" # Kode spesifik untuk Flood Wait
    except RPCError as e:
        error_msg = str(e).lower()
        if "user_already_participant" in error_msg:
            return False, "ALREADY_MEMBER" # Kode untuk sudah menjadi anggota
        elif "privacy" in error_msg or "restricted" in error_msg or "mutual contact" in error_msg:
            return False, "PRIVACY_RESTRICTED" # Kode terpadu untuk semua jenis privasi
        elif "users_too_much" in error_msg: # Error saat akun sudah terlalu banyak mengundang
            return False, "INVITE_LIMIT_REACHED" # Kode untuk limit undangan akun
        elif "banned from sending messages" in error_msg:
            return False, "BANNED_IN_SUPERGROUP" # Kode untuk akun di-ban dari grup
        elif "could not find the input entity" in error_msg:
            return False, "ENTITY_NOT_FOUND"
        elif "chat_admin_required" in error_msg:
            return False, "ADMIN_REQUIRED"
        elif "user_banned_in_channel" in error_msg:
            return False, "USER_IS_BANNED"
        elif "chat_full" in error_msg:
            return False, "GROUP_IS_FULL"
        else:
            return False, f"RPC_ERROR:{e}" # Kode untuk error RPC lainnya
    except Exception as e:
        if "could not find the input entity" in str(e).lower():
            return False, "ENTITY_NOT_FOUND"
        return False, f"GENERAL_ERROR:{e}" # Kode untuk error umum

async def send_group_link(user_client, user_id, user_username, target_entity, custom_invite_link=None):
    """Mengirim link undangan grup ke user via DM."""
    group_link = custom_invite_link
    if not group_link:
        if hasattr(target_entity, 'username') and target_entity.username:
            group_link = f"https://t.me/{target_entity.username}"
        else:
            return False, "Grup privat & tidak ada link undangan yang diberikan."
    
    try:
        # Coba dapatkan entitas user. Ini penting jika akun belum "mengenal" user.
        user_entity = None
        try:
            # Cara paling cepat adalah via ID jika sudah ada di cache sesi.
            user_entity = await user_client.get_entity(user_id)
        except ValueError:
            # Jika gagal (tidak ada di cache), coba cari via username jika tersedia.
            if user_username and user_username != 'N/A':
                print(f"[INFO] Gagal dapatkan user {user_id} via ID, mencoba via username @{user_username}...")
                user_entity = await user_client.get_entity(user_username)
        
        if not user_entity:
            return False, "Gagal menemukan user (tidak ada di cache & tidak ada username valid)"

        text = (
            "Halo! Kami mengundang Anda untuk bergabung dengan grup kami.\n\n"
            "Silakan klik link di bawah ini untuk bergabung:\n"
            f"➡️ [Gabung Grup]({group_link})"
        )
        await user_client.send_message(user_entity, text, parse_mode='md')
        return True, "Link terkirim"
    except (ValueError, TypeError):
        return False, "Gagal menemukan user via ID atau Username"
    except (UserPrivacyRestrictedError, UserNotMutualContactError):
        return False, "Privasi user/Bukan kontak mutual"
    except Exception as e:
        return False, f"Error kirim link: {e}"

async def run_broadcast(event, user_client, session_name, target_str, delay_minutes, invite_link, member_list, source_filename, mode='default', skip_user_ids=None, max_users_per_session=None):
    """Fungsi utama untuk menjalankan proses broadcast/add member."""
    TASK_STATE[session_name] = {
        "running": True,
        "task_name": "broadcast",
        "stop_requested": False,
    }

    # Inisialisasi set skip jika tidak disediakan
    if skip_user_ids is None:
        skip_user_ids = set()

    stats = {'processed': 0, 'added': 0, 'link_sent': 0, 'failed': 0, 'already_member': 0, 'skipped_privacy': 0}
    history_log = []
    start_time = datetime.now()
    status_message = await event.reply(f"Memulai proses broadcast dengan akun `{session_name}`...")

    stop_reason_code = 'error' # Default stop reason
    try:
        await user_client.connect()
        if not await user_client.is_user_authorized():
            await event.reply(f"❌ Gagal otorisasi dengan akun `{session_name}`. Mungkin perlu login ulang.")
            return

        # 1. Dapatkan entitas grup tujuan
        try:
            # Coba konversi ke integer dulu, karena ID grup bisa negatif
            try:
                target_id = int(target_str)
                target_entity = await user_client.get_entity(target_id)
            except ValueError:
                # Jika bukan integer, anggap sebagai username (cth: @namagrup)
                target_entity = await user_client.get_entity(target_str)
        except (ValueError, TypeError, Exception) as e:
            error_text = str(e)
            if "Cannot find any entity" in error_text:
                await event.reply(
                    f"❌ **Gagal Menemukan Grup Target!**\n\n"
                    f"Akun `{session_name}` tidak dapat menemukan grup `{target_str}`.\n\n"
                    f"**Penyebab paling umum:** Akun `{session_name}` **belum bergabung** dengan grup target tersebut. Untuk grup privat (dengan ID seperti `-100...`), sebuah akun wajib menjadi anggota untuk dapat berinteraksi dengannya.\n\n"
                    f"**Solusi:** Pastikan akun `{session_name}` sudah menjadi anggota grup target, lalu coba lagi."
                )
            else:
                await event.reply(f"❌ Gagal menemukan grup target `{target_str}`. Error: {e}")
            return

        # PENGECEKAN BARU: Jika grupnya privat, link undangan wajib ada.
        # Ini mencegah bot berjalan sia-sia jika konfigurasi salah.
        is_private_group = not hasattr(target_entity, 'username') or not target_entity.username
        if is_private_group and not invite_link and mode == 'default':
            await event.reply(
                f"❌ **Kesalahan Konfigurasi!**\n\n"
                f"Grup target **{target_entity.title}** adalah grup **privat**. "
                f"Untuk mengirim undangan ke anggota yang membatasi privasi, Anda **wajib** menyertakan link undangan di dalam perintah.\n\n"
                f"Contoh: `/addgrup {session_name} {target_str} {delay_minutes} https://t.me/joinchat/LINK_ANDA`"
            )
            return

        # Pre-fetch existing members to avoid re-adding
        await status_message.edit(f"⏳ Mengambil daftar anggota yang sudah ada di grup **{target_entity.title}**...\nIni bisa memakan waktu untuk grup besar.")
        existing_member_ids = set()
        try:
            # Hanya admin yang bisa melihat semua anggota di grup privat.
            # Jika akun user bukan admin, ini mungkin tidak mengembalikan semua anggota.
            async for member in user_client.iter_participants(target_entity):
                existing_member_ids.add(member.id)
            
            await event.reply(f"✅ Ditemukan **{len(existing_member_ids)}** anggota di grup target. Mereka akan dilewati secara otomatis.")
        except Exception as e:
            await event.reply(f"⚠️ Gagal mengambil daftar anggota grup target. Pengecekan duplikat mungkin tidak akurat. Error: {e}")
            # Tetap lanjutkan, bot akan mengandalkan error 'user_already_participant'

        # 2. Gunakan daftar anggota yang sudah difilter dari pool manager
        all_members = member_list
        if not all_members:
            await event.reply(f"ℹ️ Akun `{session_name}` tidak menerima daftar anggota untuk diproses. Mungkin semua sudah diproses atau terfilter.")
            # Ini bukan error, jadi kita kembalikan 'completed'
            stop_reason_code = 'completed'
            return
        
        total_users = len(all_members)
        await status_message.edit(
            f"🎯 Grup Target: **{target_entity.title}**\n"
            f"📂 File Scrape: `{source_filename}`\n"
            f"👥 Total User: **{total_users}**\n"
            f"⏳ Jeda: **{delay_minutes} menit**\n\n"
            "Memulai proses..."
        )

        # 4. Mulai loop broadcast
        last_status_text = "Menunggu user pertama..."
        for i, (uid, username, name) in enumerate(all_members, 1):
            if TASK_STATE.get(session_name, {}).get("stop_requested"):
                await event.reply("⏹️ Proses broadcast dihentikan oleh pengguna.")
                break

            stats['processed'] = i
            current_user_display = f"`{name}` (ID: `{uid}`)"
            
            sleep_after_action = False # Flag untuk menentukan apakah perlu jeda
            status_code = ""
            status_detail = ""

            # Lewati user jika sudah diproses oleh akun lain di pool yang sama
            if uid in skip_user_ids:
                continue

            # Pengecekan utama: Lewati jika user sudah ada di daftar anggota yang diambil sebelumnya.
            if uid in existing_member_ids:
                stats['already_member'] += 1
                last_status_text = f"👥 {current_user_display}: Sudah menjadi anggota (dilewati)."
                status_code = "already_member"
                status_detail = "User already in group (pre-check)."
            else:
                should_process = True
                try:
                    user_to_add = await user_client.get_entity(uid)
                    if user_to_add.bot:
                        stats['failed'] += 1
                        last_status_text = f"⏭️ {current_user_display}: Dilewati (akun bot)."
                        should_process = False
                        status_code = "skipped_bot"
                        status_detail = "User is a bot."
                except ValueError:
                    # Error ini sering terjadi jika sesi saat ini belum "melihat" user (tidak ada di grup yang sama).
                    # Kita akan melewati pengecekan bot dan membiarkan upaya penambahan yang menentukan.
                    print(f"[INFO] Tidak dapat menemukan info user {uid} (kemungkinan tidak ada di grup yang sama). Melanjutkan untuk mencoba menambahkan...")
                    should_process = True # Tetap lanjutkan proses
                except Exception as e:
                    stats['failed'] += 1
                    last_status_text = f"❌ {current_user_display}: Gagal dapatkan info user ({e})."
                    should_process = False
                    status_code = "failed"
                    status_detail = f"Failed to get user entity: {e}"

                if should_process:
                    success, reason_code = await add_user_to_group(user_client, target_entity, uid, username)

                    if success:
                        stats['added'] += 1
                        # PERUBAHAN: Gunakan reason_code untuk status yang lebih akurat
                        last_status_text = f"✅ {current_user_display}: {reason_code}."
                        sleep_after_action = True
                        existing_member_ids.add(uid) # Tambahkan ke set agar tidak diproses lagi
                        status_code = "added" # Kode internal tetap 'added' untuk konsistensi riwayat
                        status_detail = f"Success: {reason_code}"
                    elif reason_code == "PRIVACY_RESTRICTED":
                        # Mode 'fast' akan melewati user dengan privasi, mode 'default' akan mencoba kirim link.
                        if mode == 'fast':
                            stats['skipped_privacy'] += 1
                            last_status_text = f"⏩ {current_user_display}: Dilewati (akun privat)."
                            status_code = "skipped_privacy"
                            status_detail = "User skipped due to privacy settings (fast mode)."
                        else:
                            # Jika gagal karena privasi (atau error lama 'cannot cast'), coba kirim link
                            link_success, link_reason = await send_group_link(user_client, uid, username, target_entity, invite_link) # type: ignore
                            if link_success:
                                stats['link_sent'] += 1
                                last_status_text = f"🔗 {current_user_display}: Gagal tambah (privasi/grup dasar), link undangan terkirim."
                                sleep_after_action = True
                                status_code = "link_sent"
                                status_detail = "Could not add due to privacy, sent invite link."
                            else:
                                stats['failed'] += 1
                                last_status_text = f"❌ {current_user_display}: Gagal tambah & gagal kirim link ({link_reason})."
                                status_code = "failed"
                                status_detail = f"Failed to add (privacy) and failed to send link: {link_reason}"
                    elif reason_code == "ALREADY_MEMBER":
                        stats['already_member'] += 1
                        last_status_text = f"👥 {current_user_display}: Sudah menjadi anggota."
                        existing_member_ids.add(uid) # Pastikan ada di set
                        status_code = "already_member"
                        status_detail = "User already in group (API response)."
                    else:
                        # Pengecekan untuk error limit dari Telegram yang harus menghentikan proses
                        if reason_code.startswith("FLOOD_WAIT:"):
                            try:
                                wait_seconds = int(reason_code.split(":")[1])
                                wait_duration_str = str(timedelta(seconds=wait_seconds))
                                await event.reply(f"🛑 **LIMIT TELEGRAM TERDETEKSI (FLOOD WAIT)!**\n\nAkun `{session_name}` telah dibatasi oleh Telegram karena terlalu banyak permintaan. Proses untuk akun ini dihentikan secara otomatis.\n\n**Rekomendasi:** Istirahatkan akun ini setidaknya selama **{wait_duration_str}**.")
                                status_detail = f"Telegram flood wait limit hit ({wait_seconds}s)."
                            except (IndexError, ValueError):
                                await event.reply(f"🛑 **LIMIT TELEGRAM TERDETEKSI (FLOOD WAIT)!**\n\nAkun `{session_name}` telah dibatasi oleh Telegram karena terlalu banyak permintaan. Proses untuk akun ini dihentikan secara otomatis.\n\n**Rekomendasi:** Istirahatkan akun ini setidaknya selama 24 jam.")
                                status_detail = "Telegram flood wait limit hit (unknown duration)."
                            
                            stop_reason_code = 'flood_wait'
                            break # Hentikan loop untuk akun ini, akan dilanjutkan oleh akun lain
                        elif reason_code == "BANNED_IN_SUPERGROUP":
                            await event.reply(f"🛑 **AKUN DI-BAN DARI GRUP!**\n\nAkun `{session_name}` sepertinya telah di-ban atau dibatasi untuk menambahkan anggota di grup target. Proses untuk akun ini dihentikan.\n\n**Rekomendasi:** Coba gunakan akun lain atau periksa status akun `{session_name}` secara manual.")
                            stats['failed'] += 1
                            status_detail = "Account is banned from inviting in the target group."
                            stop_reason_code = 'banned'
                            break # Hentikan loop untuk akun ini
                        elif reason_code == "INVITE_LIMIT_REACHED" or "too many requests" in reason_code.lower():
                            await event.reply(f"🛑 **LIMIT TELEGRAM TERDETEKSI!**\n\nAkun `{session_name}` telah dibatasi oleh Telegram karena terlalu banyak permintaan. Proses untuk akun ini dihentikan secara otomatis.\n\n**Rekomendasi:** Istirahatkan akun ini setidaknya selama 24 jam.")
                            last_status_text = f"🛑 {current_user_display}: Gagal (LIMIT TERCAPAI)."
                            stats['failed'] += 1
                            status_code = "failed"
                            status_detail = "Account invite limit reached or too many requests."
                            stop_reason_code = 'flood_wait' # Perlakukan sebagai flood wait agar pool beralih akun
                            break # Hentikan loop untuk akun ini
                        elif reason_code == "ENTITY_NOT_FOUND":
                            stats['failed'] += 1
                            last_status_text = f"❓ {current_user_display}: Gagal (User tidak dikenal)."
                            status_code = "failed"
                            status_detail = "Entity not found. The adding account may not know this user (not in contacts, no mutual groups)."
                        elif reason_code == "USER_IS_BANNED":
                            stats['failed'] += 1
                            last_status_text = f"🚫 {current_user_display}: Gagal (User di-ban dari grup)."
                            status_code = "failed"
                            status_detail = "The target user is banned in the destination group."
                        elif reason_code == "ADMIN_REQUIRED":
                            await event.reply(f"🛑 **HAK ADMIN DIPERLUKAN!**\n\nAkun `{session_name}` tidak memiliki hak admin untuk mengundang anggota ke grup target. Proses untuk akun ini dihentikan.")
                            status_detail = "The account lacks admin privileges to invite users."
                            stop_reason_code = 'admin_required'
                            break
                        elif reason_code == "GROUP_IS_FULL":
                            await event.reply(f"🛑 **GRUP PENUH!**\n\nTarget grup **{target_entity.title}** sudah penuh. Tidak bisa menambahkan anggota lagi. Proses dihentikan.")
                            status_detail = "The target group is full."
                            stop_reason_code = 'group_full'
                            break
                        else:
                            stats['failed'] += 1
                            last_status_text = f"❌ {current_user_display}: Gagal ({reason_code})."
                            status_code = "failed"
                            status_detail = f"Failed to add: {reason_code}"

            # Simpan log riwayat
            history_log.append({
                "timestamp": datetime.now().isoformat(),
                "user_id": uid,
                "user_name": name,
                "user_username": username,
                "target_group_id": target_entity.id,
                "target_group_title": target_entity.title,
                "status": status_code,
                "details": status_detail,
            })

            # Tampilkan status di konsol untuk analisis
            print(f"[BROADCAST] [{session_name}] {i}/{total_users} - {last_status_text}")

            # Update status message setelah setiap user
            elapsed_time = datetime.now() - start_time
            summary_text = (
                f"🔄 **Broadcast Berjalan...** ({i}/{total_users})\n\n"
                f"**Status Terakhir:**\n{last_status_text}\n\n"
                f"--- **Statistik Total** ---\n"
                f"✅ **Sukses (Tambah/Undang):** {stats['added']}\n"
                f"🔗 **Link Terkirim:** {stats['link_sent']}\n"
                f"⏩ **Dilewati (Privasi):** {stats['skipped_privacy']}\n"
                f"👥 **Sudah Jadi Anggota:** {stats['already_member']}\n"
                f"❌ **Gagal:** {stats['failed']} (termasuk bot & error)\n\n"
                f"⏱️ **Durasi:** {str(elapsed_time).split('.')[0]}"
            )
            try:
                await status_message.edit(summary_text)
            except FloodWaitError as fwe:
                await asyncio.sleep(fwe.seconds + 5)
            except RPCError:
                pass

            # PERUBAHAN BARU: Cek batas proses per sesi
            if max_users_per_session is not None and stats['processed'] >= max_users_per_session:
                # Tidak perlu kirim pesan di sini, ditangani oleh pool manager
                print(f"[INFO] Akun `{session_name}` telah mencapai batas proses ({max_users_per_session} user).")
                stop_reason_code = 'daily_limit_reached'
                break

            # --- JEDA ANTAR-USER (MODE BATCH) ---
            # Jeda panjang (delay_minutes) sekarang berfungsi sebagai jeda antar-akun.
            # Di sini kita gunakan jeda singkat acak untuk membuat aktivitas lebih natural.
            if i < total_users: # Jangan jeda setelah user terakhir.
                short_delay = random.randint(10, 25) # Jeda acak antara 10 dan 25 detik
                await asyncio.sleep(short_delay)

        # 5. Kirim laporan akhir
        # Laporan akhir per-sesi tidak lagi dikirim di sini untuk menghindari kebingungan.
        # Laporan gabungan akan dikirim oleh run_pooled_broadcast_task di akhir.
        # Tentukan alasan berhenti yang benar untuk dikembalikan ke pool manager.
        if TASK_STATE.get(session_name, {}).get("stop_requested"):
            stop_reason_code = 'stopped_by_user'
        # Jika loop selesai tanpa error, tandai sebagai 'completed'.
        # 'error' adalah nilai awal, jadi jika tidak berubah, berarti loop selesai.
        elif stop_reason_code == 'error':
            stop_reason_code = 'completed'

    except Exception as e:
        await event.reply(f"❌ Terjadi error saat broadcast dengan `{session_name}`:\n`{e}`")
        traceback.print_exc()
        stop_reason_code = 'error'
    finally:
        # Jangan simpan file di sini. Kembalikan log ke pool manager.
        processed_ids_this_run = {log.get("user_id") for log in history_log}

        if session_name in TASK_STATE:
            del TASK_STATE[session_name]
        if user_client.is_connected():
            await user_client.disconnect()
        
    # Kembalikan log agar bisa digabungkan oleh pool manager
    return stop_reason_code, processed_ids_this_run, stats, history_log

async def scrape_group_members(user_client, group_entity):
    """Scrape anggota dari satu grup."""
    members_dict = {}
    try:
        async for user in user_client.iter_participants(group_entity, limit=None):
            if isinstance(user, User) and not user.bot:
                members_dict[user.id] = user
    except Exception:
        # Jika metode standar gagal, coba metode lain (misal: riwayat pesan)
        try:
            async for message in user_client.iter_messages(group_entity, limit=500):
                if message.sender and isinstance(message.sender, User) and not message.sender.bot:
                    if message.sender.id not in members_dict:
                        members_dict[message.sender.id] = message.sender
        except Exception:
            return False, {} # Gagal total

    return True, members_dict

async def run_scraping(event, user_client, session_name):
    """Fungsi utama untuk menjalankan proses scraping."""
    TASK_STATE[session_name] = {
        "running": True,
        "task_name": "scraper",
        "stop_requested": False,
    }

    status_message = await event.reply(f"Memulai proses scraping dengan akun `{session_name}`, ini mungkin memakan waktu...")
    start_time = datetime.now()

    groups_results = []
    groups_failed = []

    try:
        await user_client.connect()
        if not await user_client.is_user_authorized():
            await event.reply(f"❌ Gagal otorisasi dengan akun `{session_name}`. Mungkin perlu login ulang.")
            return

        dialogs = await user_client.get_dialogs()
        groups = [d.entity for d in dialogs if isinstance(d.entity, (Chat, Channel))]
        total_groups = len(groups)

        for i, group in enumerate(groups, 1):
            if TASK_STATE.get(session_name, {}).get("stop_requested"):
                await event.reply("⏹️ Proses scraping dihentikan oleh pengguna.")
                break

            await status_message.edit(f"🔄 Scraping... ({i}/{total_groups})\n\nSedang memproses: **{group.title}**")

            success, members_dict = await scrape_group_members(user_client, group)

            if success and members_dict:
                groups_results.append({
                    'group_id': group.id,
                    'group_title': group.title,
                    'group_type': 'Supergroup' if isinstance(group, Channel) else 'Basic Group',
                    'member_count': len(members_dict),
                    'members': [
                        (uid, user.username or 'N/A', user.first_name or '(No Name)')
                        for uid, user in members_dict.items()
                    ]
                })
            else:
                groups_failed.append({
                    'group_id': group.id,
                    'group_title': group.title,
                    'reason': 'Anggota tersembunyi atau tidak ada akses'
                })

        # Simpan hasil ke file Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Scraped Members"
        ws.append(['group_id', 'group_title', 'uid', 'username', 'name'])

        seen_uids = set()  # Set untuk melacak UID yang sudah ditambahkan untuk de-duplikasi
        for group in groups_results:
            group_id = group['group_id']
            group_title = group['group_title']
            for uid, username, name in group['members']:
                if uid not in seen_uids:
                    ws.append([group_id, group_title, uid, username, name])
                    seen_uids.add(uid)

        # Nama file output sesuai dengan nama sesi untuk memudahkan identifikasi dan menimpa file lama.
        output_file = SCRIPT_DIR / f"hasil_scraper_{session_name}.xlsx"
        wb.save(output_file)

        # Kirim ringkasan dan file ke user
        total_unique_members = len(seen_uids)
        elapsed_time = datetime.now() - start_time
        summary_text = (
            f"🏁 **Scraping Selesai!**\n\n"
            f"✅ **Grup Berhasil Di-scrape:** {len(groups_results)}\n"
            f"❌ **Grup Gagal/Privat:** {len(groups_failed)}\n"
            f"👥 **Total Anggota Unik Ditemukan:** {total_unique_members}\n\n"
            f"⏱️ **Durasi:** {str(elapsed_time).split('.')[0]}\n\n"
            f"Laporan lengkap disimpan dalam file Excel `{output_file}`.\nAnggota duplikat dari berbagai grup telah dihapus, hanya menyisakan entri unik."
        )
        await event.client.send_file(
            event.chat_id,
            output_file,
            caption=summary_text,
            reply_to=event.message.id,
            force_document=True,
            attributes=[DocumentAttributeFilename(file_name=Path(output_file).name)]
        )
        # os.remove(output_file) # File tidak lagi dihapus dan akan tersimpan di server/lokal.

    except Exception as e:
        await event.reply(f"❌ Terjadi error saat scraping dengan `{session_name}`:\n`{e}`")
        traceback.print_exc()
    finally:
        if session_name in TASK_STATE:
            del TASK_STATE[session_name]
        if user_client.is_connected():
            await user_client.disconnect()

async def run_single_group_scraping(event, user_client, session_name, source_group_str, target_group_str=None):
    """Fungsi utama untuk menjalankan proses scraping dari satu grup spesifik."""
    TASK_STATE[session_name] = {
        "running": True,
        "task_name": "scrapegrup",
        "stop_requested": False,
    }

    status_message = await event.reply(f"Memulai proses scraping grup tunggal dengan akun `{session_name}`...")
    start_time = datetime.now()

    try:
        await user_client.connect()
        if not await user_client.is_user_authorized():
            await event.reply(f"❌ Gagal otorisasi dengan akun `{session_name}`. Mungkin perlu login ulang.")
            return

        # 1. Dapatkan entitas grup target
        try:
            try:
                source_id = int(source_group_str)
                source_entity = await user_client.get_entity(source_id)
            except ValueError:
                source_entity = await user_client.get_entity(source_group_str)
        except (ValueError, TypeError, Exception) as e:
            await event.reply(f"❌ Gagal menemukan grup sumber `{source_group_str}`. Pastikan akun `{session_name}` adalah anggota grup tersebut. Error: {e}")
            return

        # FITUR BARU: Proses grup target untuk filtering
        target_group_members = set()
        skipped_due_to_target = 0
        target_entity = None
        if target_group_str:
            await status_message.edit(f"⏳ Mencari grup target filter `{target_group_str}`...")
            try:
                try:
                    target_id = int(target_group_str)
                    target_entity = await user_client.get_entity(target_id)
                except ValueError:
                    target_entity = await user_client.get_entity(target_group_str)
                
                await status_message.edit(f"⏳ Mengambil daftar anggota dari grup target **{target_entity.title}** untuk perbandingan...")
                async for member in user_client.iter_participants(target_entity):
                    target_group_members.add(member.id)
                
                if target_group_members:
                    await event.reply(f"✅ Ditemukan **{len(target_group_members)}** anggota di grup target. Anggota yang sama akan dilewati dari hasil scrape.")
                else:
                    await event.reply(f"ℹ️ Tidak ada anggota yang ditemukan di grup target **{target_entity.title}** atau akses terbatas. Scraping akan berjalan normal.")

            except Exception as e:
                await event.reply(f"⚠️ Gagal memproses grup target `{target_group_str}`. Scraping akan dilanjutkan tanpa memfilter anggota. Error: {e}")
                target_group_members = set() # Reset jika gagal


        # Fitur Baru: Coba bergabung ke grup/channel secara otomatis.
        # Ini akan gagal jika grup privat atau chat biasa, dan itu tidak masalah.
        # Jika sudah menjadi anggota, tidak akan terjadi apa-apa.
        try:
            from telethon.tl.functions.channels import JoinChannelRequest
            await user_client(JoinChannelRequest(source_entity))
            await status_message.edit(f"✅ Akun `{session_name}` mencoba bergabung/memastikan keanggotaan di **{source_entity.title}**...")
            await asyncio.sleep(2) # Jeda singkat agar status terbaca
        except Exception:
            # Abaikan error di sini (misal: jika ini grup dasar/chat privat), proses scraping akan tetap dicoba.
            pass

        # 2. Lakukan scraping
        await status_message.edit(f"🔄 Scraping... Sedang memproses: **{source_entity.title}**")
        
        success, members_dict = await scrape_group_members(user_client, source_entity)

        if not success or not members_dict:
            await event.reply(f"❌ Gagal melakukan scrape anggota dari grup **{source_entity.title}**. Kemungkinan anggota grup tersembunyi atau tidak ada akses.")
            return

        # 3. Simpan hasil ke file Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Scraped Members"
        ws.append(['group_id', 'group_title', 'uid', 'username', 'name'])
        
        final_members_to_add = {}
        for uid, user in members_dict.items():
            if uid in target_group_members:
                skipped_due_to_target += 1
                continue # Lewati jika anggota sudah ada di grup target
            final_members_to_add[uid] = user

        for uid, user in final_members_to_add.items():
            ws.append([source_entity.id, source_entity.title, uid, user.username or 'N/A', user.first_name or '(No Name)'])

        safe_group_title = "".join(c for c in source_entity.title if c.isalnum() or c in (' ', '_')).rstrip().replace(" ", "_")
        output_file = SCRIPT_DIR / f"hasil_scraper_{session_name}_{safe_group_title}.xlsx"
        wb.save(output_file)

        # 4. Kirim ringkasan dan file ke user
        total_found = len(members_dict)
        total_added_to_file = len(final_members_to_add)
        elapsed_time = datetime.now() - start_time
        
        summary_text = (
            f"🏁 **Scraping Grup Selesai!**\n\n"
            f"**Grup Sumber:** {source_entity.title}\n"
            f"👥 **Total Anggota Ditemukan:** {total_found}\n"
        )
        if target_group_str and target_entity:
            summary_text += (
                f"**Grup Target (Filter):** {target_entity.title}\n"
                f"⏭️ **Dilewati (Sudah Join):** {skipped_due_to_target}\n"
                f"✅ **Anggota Unik Ditambahkan ke File:** {total_added_to_file}\n\n"
            )
        
        summary_text += f"⏱️ **Durasi:** {str(elapsed_time).split('.')[0]}\n\nLaporan lengkap disimpan dalam file Excel `{output_file}`."

        await event.client.send_file(
            event.chat_id,
            output_file,
            caption=summary_text,
            reply_to=event.message.id,
            force_document=True,
            attributes=[DocumentAttributeFilename(file_name=Path(output_file).name)]
        )

    except Exception as e:
        await event.reply(f"❌ Terjadi error saat scraping dengan `{session_name}`:\n`{e}`")
        traceback.print_exc()
    finally:
        if session_name in TASK_STATE:
            del TASK_STATE[session_name]
        if user_client.is_connected():
            await user_client.disconnect()

# =====================================================
# BOT EVENT HANDLERS
# =====================================================

@bot_client.on(events.NewMessage(pattern='/start'))
async def start_handler(event):
    print(f"[INFO] Perintah '{event.raw_text}' dari user {event.sender_id} di chat {event.chat_id}")
    await event.reply(
        "👋 **Selamat Datang di Bot Scraper & Broadcaster!**\n\n"
        # "Saya adalah bot untuk scrape anggota grup dan melakukan broadcast.\n"
        "Gunakan /help untuk melihat daftar perintah yang tersedia."
    )

@bot_client.on(events.NewMessage(pattern='/help'))
async def help_handler(event):
    print(f"[INFO] Perintah '{event.raw_text}' dari user {event.sender_id} di chat {event.chat_id}")

    help_text_base = """**BANTUAN PENGGUNAAN BOT**

Berikut adalah format dan contoh perintah yang tersedia.

---
**PERINTAH UTAMA**
---
` /scraper <nama_sesi> `
*Fungsi:* Scrape anggota dari semua grup di akun target.
*Contoh:* `/scraper akun1`\n
` /scrapergrup <nama_sesi> <grup_sumber> [grup_target_filter] `
*Fungsi:* Scrape anggota dari satu grup sumber. Jika `grup_target_filter` diberikan, anggota yang sudah ada di grup target akan dikecualikan dari hasil.
*Contoh 1 (scrape saja):* `/scrapergrup akun1 @grup_sumber`
*Contoh 2 (scrape & filter):* `/scrapergrup akun1 @grup_sumber -100123456`

` /addgrup <nama_sesi> <target> <jeda_menit> [link_opsional] ` (Bisa multi-akun: `akun1,akun2`)
*Fungsi:* Menambah anggota dari file scrape **terbaru** untuk sesi tersebut (hasil dari /scraper atau /scrapergrup).
*Contoh:* `/addgrup akun1 @grupkeren 10`
*Contoh 2:* `/addgrup akun2 -100123456 5 https://t.me/joinchat/ABC... limit=20`

` /addgrupfast <nama_sesi> <target> <jeda_menit> ` (Bisa multi-akun: `akun1,akun2`)
*Fungsi:* Sama seperti /addgrup, tapi **melewati** anggota dengan akun privat (tidak mengirim link). Berguna untuk menambah anggota secara cepat.
*Contoh:* `/addgrupfast akun1 @grupkeren 5`

` /addgrupexcel <nama_sesi> <target> <jeda_menit> [link_opsional] ` (Bisa multi-akun: `akun1,akun2`)
*Fungsi:* Menambah anggota dengan mengunggah file Excel manual.
*Contoh:* `/addgrupexcel akun1 @grupkeren 10`

---
**OPSI TAMBAHAN**
---
` limit=<angka> `
*Fungsi:* Dapat ditambahkan di akhir perintah `/addgrup`, `/addgrupfast`, dan `/addgrupexcel` untuk membatasi jumlah user yang diproses per akun dalam satu tugas. Sangat berguna untuk "pemanasan" akun baru.
*Contoh:* `/addgrupfast akun1,akun2 -100123... 10 limit=50`\n
` daily `
*Fungsi:* Jika ditambahkan, tugas akan otomatis berhenti ketika semua akun mencapai limit harian dan akan dilanjutkan kembali keesokan harinya sampai semua daftar pengguna selesai. Harus digunakan bersama dengan `limit`.
*Contoh:* `/addgrupfast akun1,akun2 -100... 10 limit=20 daily`
` file=<nama_file.xlsx> `
*Fungsi:* Menggunakan file hasil scrape dengan nama spesifik, daripada yang terbaru secara otomatis.
*Contoh:* `/addgrupfast akun1 -100... 10 file=hasil_scraper_newsb5_CUAN_DARI_RUMAH_by_DRP.xlsx`
---
**MANAJEMEN RIWAYAT (HISTORY)**
---
` history=<nama_file.xlsx> `
*Fungsi:* Menggunakan atau membuat file riwayat dengan nama spesifik. Jika tidak diberikan, `global_broadcast_history.xlsx` akan digunakan.
*Contoh:* `/addgrup akun1 -100... 10 history=proyek_A.xlsx`

` retry=failed `
*Fungsi:* Menjalankan ulang proses **hanya** untuk anggota yang tercatat 'gagal' di file riwayat. Berguna untuk mencoba kembali setelah memperbaiki masalah.
*Contoh:* `/addgrup akun1 -100... 10 retry=failed`
---
**UTILITAS**
---
` /idgrup `
*Fungsi:* Menampilkan ID dari grup saat ini.

` /status `
*Fungsi:* Melihat status semua proses yang sedang berjalan.
"""

    help_text_admin_extra = """
---
**PERINTAH ADMIN**
---
` /login <nama_sesi> `
*Contoh:* `/login akun_baru`

` /logout <nama_sesi> `
*Contoh:* `/logout akun_lama`

` /accounts `
*Fungsi:* Menampilkan semua akun yang tersimpan.

` /stop [nama_sesi] `
*Fungsi:* Menghentikan proses yang berjalan. Jika `nama_sesi` diberikan, hanya tugas untuk sesi itu yang dihentikan. Jika tidak, semua tugas akan dihentikan.
*Contoh 1 (spesifik):* `/stop akun1`
*Contoh 2 (semua):* `/stop`

` /leavegroup `
*Fungsi:* Memerintahkan bot keluar dari grup ini.
"""
    final_help_text = help_text_base
    if event.sender_id == ADMIN_ID:
        final_help_text += help_text_admin_extra

    await event.reply(final_help_text, parse_mode='md')

@bot_client.on(events.NewMessage(pattern=r'/login (\w+)'))
async def login_handler(event):
    print(f"[INFO] Perintah '{event.raw_text}' dari user {event.sender_id} di chat {event.chat_id}")
    # Check if user is admin
    if event.sender_id != ADMIN_ID:
        await event.reply("❌ Anda tidak memiliki izin untuk menggunakan perintah ini.")
        print(f"[WARNING] Unauthorized login attempt from {event.sender_id}")
        return
    
    session_name = event.pattern_match.group(1)
    session_path = Path(SESSIONS_DIR) / f"{session_name}.session"
    if session_path.exists():
        await event.reply(f"⚠️ Sesi dengan nama `{session_name}` sudah ada. Gunakan `/logout {session_name}` terlebih dahulu jika ingin login ulang.")
        return

    # Buat instance klien, tetapi jangan hubungkan dulu.
    temp_client = TelegramClient(str(session_path), API_ID, API_HASH)
    try:
        # Beri tahu admin di grup bahwa proses akan dilanjutkan di DM
        try:
            await bot_client.send_message(ADMIN_ID, f"Memulai proses login untuk sesi `{session_name}`...")
            await event.reply("✅ Perintah diterima. Silakan periksa chat pribadi Anda untuk melanjutkan proses login.")
        except Exception as e:
            await event.reply(f"❌ Gagal memulai percakapan pribadi dengan Anda (Admin). Pastikan Anda tidak memblokir bot ini.\nError: `{e}`")
            return

        print(f"[DEBUG] login_handler: Memulai untuk sesi '{session_name}' di DM admin.")

        # Gunakan loop untuk menangani respons dengan lebih robust di chat pribadi admin
        async with bot_client.conversation(ADMIN_ID, timeout=300) as conv:
            # Langkah 1: Dapatkan nomor telepon dari pengguna
            await conv.send_message("📱 Silakan masukkan nomor telepon Anda (format internasional, cth: `+628123456789`):")
            print("[DEBUG] login_handler: Menunggu nomor telepon dari user...")
            
            try:
                phone_response = await asyncio.wait_for(conv.get_response(), timeout=120)
            except asyncio.TimeoutError:
                await conv.send_message("⏱️ Waktu tunggu untuk nomor telepon habis. Coba lagi dengan /login")
                await event.reply("❌ Proses login dibatalkan karena waktu tunggu habis di chat pribadi.")
                return
            
            phone_number = phone_response.text.strip()
            print(f"[DEBUG] login_handler: Nomor diterima: {phone_number}")
            await conv.send_message(f"✅ Nomor diterima: `{phone_number}`\n⏳ Mengirim kode verifikasi...")
            
            # Langkah 2: Hubungkan, kirim kode
            try:
                print("[DEBUG] login_handler: Menghubungkan temporary client untuk mengirim kode...")
                await temp_client.connect()
                print("[DEBUG] login_handler: Temporary client terhubung.")
                
                code_req = await asyncio.wait_for(
                    temp_client.send_code_request(phone_number),
                    timeout=60.0
                )
                phone_code_hash = code_req.phone_code_hash
                print("[DEBUG] login_handler: Permintaan kode berhasil dikirim.")
                await conv.send_message("✅ Kode verifikasi telah dikirim ke nomor Anda.")
                
            except asyncio.TimeoutError:
                print(f"[ERROR] login_handler: Timeout saat mengirim kode")
                await conv.send_message(f"⏱️ Timeout saat mengirim kode. Coba lagi nanti.")
                await event.reply("❌ Proses login gagal (timeout saat kirim kode).")
                return
            except FloodWaitError as fwe:
                print(f"[ERROR] login_handler: Flood wait error: {fwe.seconds} detik")
                await conv.send_message(f"⚠️ Terlalu banyak percobaan. Tunggu {fwe.seconds} detik sebelum mencoba lagi.")
                await event.reply(f"❌ Proses login gagal (terkena Flood Wait).")
                return
            except Exception as e:
                print(f"[ERROR] login_handler: Gagal mengirim kode: {e}")
                traceback.print_exc()
                await conv.send_message(f"❌ Gagal mengirim kode:\n`{str(e)}`\n\nCoba lagi nanti atau hubungi support.")
                await event.reply(f"❌ Proses login gagal. Detail error ada di chat pribadi.")
                return
            finally:
                if temp_client.is_connected():
                    await temp_client.disconnect()
                    print("[DEBUG] login_handler: Temporary client diputuskan setelah mengirim kode.")

            # Langkah 3: Dapatkan kode verifikasi dari pengguna
            await conv.send_message("📝 Silakan masukkan kode verifikasi yang diterima:")
            print("[DEBUG] login_handler: Menunggu kode verifikasi dari user...")
            
            try:
                code_response = await asyncio.wait_for(conv.get_response(), timeout=300)
            except asyncio.TimeoutError:
                await conv.send_message("⏱️ Waktu tunggu untuk kode verifikasi habis. Coba lagi dengan /login")
                await event.reply("❌ Proses login dibatalkan karena waktu tunggu habis di chat pribadi.")
                return
                
            verification_code = code_response.text.strip()
            print("[DEBUG] login_handler: Kode diterima.")

            # Langkah 4: Hubungkan kembali dan coba untuk sign-in
            try:
                print("[DEBUG] login_handler: Menghubungkan kembali temporary client untuk sign-in...")
                await temp_client.connect()
                print("[DEBUG] login_handler: Connected untuk sign-in.")
                
                try:
                    await asyncio.wait_for(
                        temp_client.sign_in(phone_number, verification_code, phone_code_hash=phone_code_hash),
                        timeout=30.0
                    )
                    print("[DEBUG] login_handler: Sign-in berhasil.")
                except SessionPasswordNeededError:
                    await conv.send_message("🔐 Akun Anda dilindungi 2FA. Silakan masukkan kata sandi (password):")
                    print("[DEBUG] login_handler: Meminta password 2FA...")
                    
                    try:
                        password_response = await asyncio.wait_for(conv.get_response(), timeout=300)
                    except asyncio.TimeoutError:
                        await conv.send_message("⏱️ Waktu tunggu untuk password habis. Coba lagi dengan /login")
                        await event.reply("❌ Proses login dibatalkan karena waktu tunggu habis di chat pribadi.")
                        return
                    
                    password = password_response.text.strip()
                    print("[DEBUG] login_handler: Password 2FA diterima. Mencoba sign-in dengan password...")
                    await asyncio.wait_for(
                        temp_client.sign_in(password=password),
                        timeout=30.0
                    )
                    print("[DEBUG] login_handler: Sign-in dengan 2FA berhasil.")
                
                me = await temp_client.get_me()
                print(f"[INFO] login_handler: Berhasil login sebagai {me.first_name}. Sesi '{session_name}' disimpan.")
                await conv.send_message(f"✅ **Login Berhasil!**\n\n👤 Nama: **{me.first_name}**\n📋 Username: `@{me.username or 'N/A'}`\n🔢 ID: `{me.id}`\n\n📱 Sesi disimpan sebagai: `{session_name}`")
                await event.reply(f"✅ Sesi `{session_name}` berhasil disimpan.")
                
            except (ValueError, RPCError) as e:
                print(f"[ERROR] login_handler: Error saat sign-in: {e}")
                await conv.send_message(f"❌ Kode atau password salah:\n`{str(e)}`")
                await event.reply(f"❌ Proses login gagal untuk sesi `{session_name}`. Detail error ada di chat pribadi.")
                return
            except asyncio.TimeoutError:
                print(f"[ERROR] login_handler: Timeout saat sign-in")
                await conv.send_message(f"⏱️ Timeout saat sign-in. Coba lagi dengan /login")
                await event.reply(f"❌ Proses login gagal untuk sesi `{session_name}` (timeout).")
                return
            except Exception as e:
                print(f"[ERROR] login_handler: Error sign-in: {e}")
                traceback.print_exc()
                await conv.send_message(f"❌ Error saat login:\n`{str(e)}`")
                await event.reply(f"❌ Proses login gagal untuk sesi `{session_name}`. Detail error ada di chat pribadi.")
                return
    except asyncio.TimeoutError:
        print("[ERROR] login_handler: Proses login timeout (300 detik).")
        # Pesan ini mungkin tidak terkirim jika conversation sudah timeout, tapi kita coba.
        # Pesan utama akan dikirim di dalam blok conversation.
        await event.reply("⏱️ Waktu login habis (total 5 menit). Silakan coba lagi dengan /login")
    except Exception as e:
        print(f"[ERROR] login_handler: Terjadi error tak terduga: {e}")
        traceback.print_exc()
        await event.reply(f"❌ Terjadi error tak terduga saat login:\n`{str(e)}`")
    finally:
        try:
            if temp_client.is_connected():
                await temp_client.disconnect()
                print(f"[DEBUG] login_handler: Temporary client untuk sesi '{session_name}' diputuskan di blok final.")
        except Exception as e:
            print(f"[WARNING] Error saat disconnect: {e}")

@bot_client.on(events.NewMessage(pattern=r'/logout (\w+)'))
async def logout_handler(event):
    print(f"[INFO] Perintah '{event.raw_text}' dari user {event.sender_id} di chat {event.chat_id}")
    # Check if user is admin
    if event.sender_id != ADMIN_ID:
        await event.reply("❌ Anda tidak memiliki izin untuk menggunakan perintah ini.")
        return
    
    session_name = event.pattern_match.group(1)
    session_path = Path(SESSIONS_DIR) / f"{session_name}.session"
    if not session_path.exists():
        await event.reply(f"❌ Sesi `{session_name}` tidak ditemukan.")
        return
    try:
        os.remove(session_path)
        journal_path = Path(SESSIONS_DIR) / f"{session_name}.session-journal"
        if journal_path.exists():
            os.remove(journal_path)
        await event.reply(f"✅ Sesi `{session_name}` berhasil dihapus (logout).")
    except Exception as e:
        await event.reply(f"❌ Gagal menghapus sesi: `{e}`")

@bot_client.on(events.NewMessage(pattern='/accounts'))
async def accounts_handler(event):
    print(f"[INFO] Perintah '{event.raw_text}' dari user {event.sender_id} di chat {event.chat_id}")
    # Check if user is admin
    if event.sender_id != ADMIN_ID:
        await event.reply("❌ Anda tidak memiliki izin untuk menggunakan perintah ini.")
        return

    # Filter sesi bot itu sendiri untuk menghindari error 'database is locked'
    session_files = [f for f in Path(SESSIONS_DIR).glob('*.session') if f.stem != BOT_SESSION_NAME]
    if not session_files: # Sekarang ini hanya akan kosong jika tidak ada sesi USER
        await event.reply("Tidak ada akun user yang tersimpan. Gunakan `/login <nama_sesi>` untuk menambahkan.")
        return

    status_message = await event.reply("⏳ Mengambil informasi akun, harap tunggu...")
    
    message_lines = ["👤 **Daftar Akun User Tersimpan:**"]
    
    for i, session_file in enumerate(session_files, 1):
        session_name = session_file.stem
        temp_client = TelegramClient(str(session_file.resolve()), API_ID, API_HASH)
        
        try:            
            async def check_account_status():
                """Fungsi helper untuk menghubungkan, memeriksa otorisasi, dan mendapatkan info user."""
                await temp_client.connect()
                if await temp_client.is_user_authorized():
                    me = await temp_client.get_me()
                    phone = f"+{me.phone}" if me.phone else "N/A"
                    username = f"@{me.username}" if me.username else "N/A"
                    return f"{i}. **{session_name}**\n   - Nama: `{me.first_name}`\n   - Username: `{username}`\n   - No. HP: `{phone}`"
                else:
                    return f"{i}. **{session_name}**\n   - Status: `Sesi tidak valid/perlu login ulang`"

            # Bungkus seluruh proses pengecekan dalam satu timeout
            result_line = await asyncio.wait_for(check_account_status(), timeout=20.0)
            message_lines.append(result_line)
        except asyncio.TimeoutError:
            print(f"[ERROR] Timeout saat memeriksa akun {session_name}")
            message_lines.append(f"{i}. **{session_name}**\n   - Status: `Timeout saat menghubungkan`")
        except Exception as e:
            print(f"[ERROR] Gagal memeriksa akun {session_name}: {e}")
            message_lines.append(f"{i}. **{session_name}**\n   - Status: `Gagal terhubung atau sesi korup`")
        finally:
            if temp_client.is_connected():
                await temp_client.disconnect()

    await status_message.edit("\n\n".join(message_lines), parse_mode='md')

@bot_client.on(events.NewMessage(pattern=r'/scraper (\w+)$'))
async def scraper_handler(event):
    print(f"[INFO] Perintah '{event.raw_text}' dari user {event.sender_id} di chat {event.chat_id}")
    session_name = event.pattern_match.group(1)
    if TASK_STATE.get(session_name, {}).get("running"):
        await event.reply(f"⚠️ Akun `{session_name}` sedang menjalankan tugas `{TASK_STATE[session_name]['task_name']}`. Harap tunggu.")
        return
    session_path = Path(SESSIONS_DIR) / f"{session_name}.session"
    if not session_path.exists():
        await event.reply(f"❌ Sesi `{session_name}` tidak ditemukan. Gunakan `/login {session_name}` atau periksa daftar dengan `/accounts`.")
        return
    user_client = TelegramClient(str(session_path), API_ID, API_HASH)
    asyncio.create_task(run_scraping(event, user_client, session_name))

@bot_client.on(events.NewMessage(pattern=r'/scrapergrup (\w+) ([^ ]+)(?:\s+(.+))?'))
async def scrapergrup_handler(event):
    print(f"[INFO] Perintah '{event.raw_text}' dari user {event.sender_id} di chat {event.chat_id}")
    session_name = event.pattern_match.group(1)
    source_group_str = event.pattern_match.group(2)
    target_group_str = event.pattern_match.group(3) # Bisa None jika tidak ada

    if TASK_STATE.get(session_name, {}).get("running"):
        await event.reply(f"⚠️ Akun `{session_name}` sedang menjalankan tugas `{TASK_STATE[session_name]['task_name']}`. Harap tunggu.")
        return
        
    session_path = Path(SESSIONS_DIR) / f"{session_name}.session"
    if not session_path.exists():
        await event.reply(f"❌ Sesi `{session_name}` tidak ditemukan. Gunakan `/login {session_name}` atau periksa daftar dengan `/accounts`.")
        return
        
    user_client = TelegramClient(str(session_path), API_ID, API_HASH)
    asyncio.create_task(run_single_group_scraping(event, user_client, session_name, source_group_str, target_group_str))

async def run_pooled_broadcast_task(event, session_names, target_str, delay_minutes, invite_link, mode, excel_file_path=None, start_message_id=None, max_users_per_session=None, daily=False, history_filename=None, retry_failed_only=False):
    """Manajer tugas yang menjalankan broadcast di beberapa akun secara berurutan."""
    globally_processed_ids = set()
    is_first_run = True
    total_stats = {'processed': 0, 'added': 0, 'link_sent': 0, 'failed': 0, 'already_member': 0, 'skipped_privacy': 0}
    accounts_used = []
    all_history_logs = [] # Kumpulkan semua log di sini
    start_time = datetime.now()

    # LOGIKA BARU: Tentukan file Excel sumber SEKALI di awal, jika tidak disediakan.
    # Ini memastikan semua akun dalam pool menggunakan file yang sama.
    source_excel_path = excel_file_path
    if not source_excel_path:
        try:
            # LOGIKA BARU: Cari file scrape APAPUN yang paling baru, tidak terikat pada sesi tertentu.
            search_pattern = "hasil_scraper_*.xlsx"
            files = [(p, p.stat().st_mtime) for p in SCRIPT_DIR.glob(search_pattern)]
            if not files:
                await event.reply(
                    f"❌ **File Scrape Tidak Ditemukan!**\n\n"
                    f"Saya tidak dapat menemukan file hasil scrape sama sekali di direktori bot.\n"
                    f"Pastikan Anda telah menjalankan `/scraper` atau `/scrapergrup` terlebih dahulu.\n\n"
                    f"(Pencarian dilakukan untuk file dengan pola: `{search_pattern}`)"
                )
                return
            files.sort(key=lambda x: x[1], reverse=True)
            source_excel_path = str(files[0][0])
            await event.reply(f"ℹ️ Menggunakan file scrape terbaru yang ditemukan: `{Path(source_excel_path).name}`")
        except Exception as e:
            await event.reply(f"❌ Terjadi error saat mencari file scrape terbaru: {e}")
            return

    # Baca file sumber ke dalam memori
    all_members = []
    try:
        wb = openpyxl.load_workbook(source_excel_path)
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(row): continue
            if len(row) >= 3 and row[2]:
                try:
                    uid = int(row[2])
                    username = row[3] if len(row) > 3 else 'N/A'
                    name = row[4] if len(row) > 4 else '(No Name)'
                    all_members.append((uid, username, name))
                except (ValueError, TypeError):
                    print(f"[WARNING] Melewati baris {i} di {Path(source_excel_path).name}: UID '{row[2]}' bukan angka yang valid.")
                    continue
    except Exception as e:
        await event.reply(f"❌ Gagal membaca file Excel sumber `{Path(source_excel_path).name}`. Error: {e}")
        return

    # Tentukan path riwayat dan filter anggota berdasarkan itu
    history_path = HISTORY_DIR / history_filename if history_filename else DEFAULT_HISTORY_FILE
    status_map, processed_ids_from_history = load_history_data(history_path)

    initial_total = len(all_members)
    if retry_failed_only:
        failed_statuses = {'failed', 'privacy_restricted', 'banned', 'invite_limit_reached', 'rpc_error', 'general_error', 'banned_in_supergroup'}
        all_members = [
            member for member in all_members 
            if status_map.get(member[0]) in failed_statuses
        ]
        await event.reply(f"🔄 **Mode Retry Gagal:** Menargetkan **{len(all_members)}** dari **{initial_total}** pengguna yang sebelumnya gagal (berdasarkan `{history_path.name}`).")
    else:
        all_members = [
            member for member in all_members 
            if member[0] not in processed_ids_from_history
        ]
        if initial_total > 0:
             await event.reply(f"ℹ️ **Filter Riwayat:** Dari **{initial_total}** pengguna, **{len(all_members)}** akan diproses. Sisanya (`{initial_total - len(all_members)}`) sudah ada di riwayat (`{history_path.name}`).")

    if not all_members:
        await event.reply("✅ Tidak ada pengguna baru untuk diproses. Semua target sudah ada di riwayat atau tidak cocok dengan filter.")
        return

    # Buat pesan status awal
    await event.reply(
        f"**Tugas Dimulai**\n\n"
        f"👥 **Total Pengguna untuk Diproses:** {len(all_members)}\n"
        f"📂 **File Riwayat:** `{history_path.name}`"
    )

    while True: # Loop utama untuk siklus harian
        for i, session_name in enumerate(session_names):
            session_name = session_name.strip() # Hapus spasi
            if not session_name: continue

            if TASK_STATE.get(session_name, {}).get("running"):
                await event.reply(f"⚠️ Melewati akun `{session_name}` karena sedang menjalankan tugas lain.")
                continue

            if not is_first_run:
                await event.reply(f"▶️ Melanjutkan tugas dengan akun berikutnya: `{session_name}`")
            is_first_run = False
            if session_name not in accounts_used:
                accounts_used.append(session_name)

            session_path = Path(SESSIONS_DIR) / f"{session_name}.session"
            if not session_path.exists():
                await event.reply(f"❌ Sesi `{session_name}` tidak ditemukan. Melewati...")
                continue

            user_client = TelegramClient(str(session_path), API_ID, API_HASH)
            
            stop_reason, processed_this_run, stats_this_run, logs_this_run = await run_broadcast(
                event, user_client, session_name, target_str, delay_minutes, 
                invite_link, member_list=all_members, source_filename=Path(source_excel_path).name,
                mode=mode, 
                skip_user_ids=globally_processed_ids,
                max_users_per_session=max_users_per_session
            )

            if logs_this_run:
                all_history_logs.extend(logs_this_run)
            if processed_this_run:
                globally_processed_ids.update(processed_this_run)
            
            if stats_this_run:
                for key in total_stats:
                    total_stats[key] += stats_this_run.get(key, 0)

            if stop_reason == 'completed':
                await event.reply(f"🎉 Semua pengguna dalam file telah diproses dengan sukses menggunakan akun `{session_name}`.")
                # Keluar dari kedua loop untuk membuat laporan akhir
                break # Keluar dari for loop
            elif stop_reason == 'stopped_by_user':
                await event.reply(f"⏹️ Tugas dihentikan oleh pengguna. Pool dihentikan.")
                return
            elif stop_reason == 'daily_limit_reached':
                if i < len(session_names) - 1:
                    await event.reply(f"ℹ️ Batas proses untuk `{session_name}` tercapai.")
                    await event.reply(f"**Jeda {delay_minutes} menit sebelum beralih ke akun berikutnya...**")
                    await asyncio.sleep(delay_minutes * 60)
                    continue
                else:
                    # Akun terakhir juga mencapai limit, akhir dari siklus hari ini
                    break # Keluar dari for loop untuk evaluasi harian
            elif stop_reason == 'flood_wait' or stop_reason == 'banned':
                if i < len(session_names) - 1:
                    if stop_reason == 'flood_wait':
                        await event.reply(f"🔁 Akun `{session_name}` terkena limit.")
                    else: # banned
                        await event.reply(f"🔁 Akun `{session_name}` di-ban dari grup.")
                    await event.reply(f"**Jeda {delay_minutes} menit sebelum beralih ke akun berikutnya...**")
                    await asyncio.sleep(delay_minutes * 60)
                    continue
                else:
                    # Akun terakhir juga terkena limit, akhir dari siklus hari ini
                    break # Keluar dari for loop untuk evaluasi harian
            elif stop_reason == 'error':
                if i < len(session_names) - 1:
                    await event.reply(f"❌ Terjadi error pada `{session_name}`. Mencoba lanjut dengan akun berikutnya...")
                    await event.reply(f"**Jeda {delay_minutes} menit sebelum beralih ke akun berikutnya...**")
                    await asyncio.sleep(delay_minutes * 60)
                    continue
                else:
                    await event.reply(f"❌ Terjadi error pada akun terakhir (`{session_name}`). Tugas dihentikan.")
                    return
        
        # --- AKHIR DARI SIKLUS (SEMUA AKUN TELAH DIGUNAKAN SEKALI) ---

        # Cek apakah tugas sudah selesai sepenuhnya
        if len(globally_processed_ids) >= len(all_members):
            print("[INFO] Tugas selesai, semua pengguna telah diproses.")
            break # Keluar dari loop `while True` untuk membuat laporan akhir

        # Jika bukan tugas harian, berhenti setelah satu siklus
        if not daily:
            print("[INFO] Tugas sekali jalan selesai.")
            break

        # Jika sampai di sini, artinya ini tugas harian dan belum selesai. Jeda sampai besok.
        now = datetime.now()
        resume_time = (now + timedelta(days=1)).replace(hour=9, minute=0, second=0, microsecond=0)
        wait_seconds = (resume_time - now).total_seconds()
        
        # Format durasi agar lebih mudah dibaca
        hours, remainder = divmod(int(wait_seconds), 3600)
        minutes, _ = divmod(remainder, 60)
        wait_duration_str = f"{hours} jam {minutes} menit"

        await event.reply(
            f"🏁 **Siklus Harian Selesai** 🏁\n\n"
            f"Semua akun telah menyelesaikan tugasnya untuk hari ini.\n"
            f"Pengguna tersisa untuk diproses: **{len(all_members) - len(globally_processed_ids)}**\n\n"
            f"Tugas akan dilanjutkan secara otomatis besok pada pukul **{resume_time.strftime('%H:%M')}** (dalam ~{wait_duration_str})."
        )
        await asyncio.sleep(wait_seconds)
        await event.reply("▶️ **Melanjutkan Tugas Harian Terjadwal...**")
        is_first_run = True # Reset agar pesan "melanjutkan dengan..." tidak muncul di akun pertama

    # Simpan semua log yang terkumpul ke file riwayat
    append_logs_to_history(history_path, all_history_logs)
    
    # Kirim laporan akhir gabungan setelah semua proses selesai
    elapsed_time = datetime.now() - start_time
    final_summary_text = (
        f"📊 **--- Laporan Akhir Gabungan ---** 📊\n\n"
        f"**Akun yang Digunakan:** {', '.join(f'`{s}`' for s in accounts_used)}\n\n"
        f"--- **Hasil Total** ---\n"
        f"✅ **Sukses (Tambah/Undang):** {total_stats['added']}\n"
        f"🔗 **Link Terkirim:** {total_stats['link_sent']}\n"
        f"⏩ **Dilewati (Privasi):** {total_stats['skipped_privacy']}\n"
        f"👥 **Sudah Jadi Anggota:** {total_stats['already_member']}\n"
        f"❌ **Gagal:** {total_stats['failed']} (termasuk bot, error, & limit)\n\n"
        f"⏱️ **Total Durasi Seluruh Tugas:** {str(elapsed_time).split('.')[0]}"
    )
    try:
        await event.reply(final_summary_text, reply_to=start_message_id)
    except Exception as e:
        print(f"[ERROR] Gagal mengirim laporan akhir gabungan: {e}")

@bot_client.on(events.NewMessage(pattern=r'/addgrup (.*)'))
async def addgrup_handler(event):
    print(f"[INFO] Perintah '{event.raw_text}' dari user {event.sender_id} di chat {event.chat_id}")
    args = [arg for arg in event.pattern_match.group(1).split(' ') if arg]

    # Parsing untuk riwayat
    history_filename = None
    history_arg = next((arg for arg in args if arg.lower().startswith('history=')), None)
    if history_arg:
        history_filename = history_arg.split('=', 1)[1]
        args.remove(history_arg)

    # Parsing untuk retry
    retry_failed_only = 'retry=failed' in [arg.lower() for arg in args]
    if retry_failed_only:
        args = [arg for arg in args if arg.lower() != 'retry=failed']

    # Parsing untuk file spesifik
    excel_file_path = None
    file_arg = next((arg for arg in args if arg.lower().startswith('file=')), None)
    if file_arg:
        specified_filename = file_arg.split('=', 1)[1]
        excel_file_path = SCRIPT_DIR / specified_filename
        if not excel_file_path.exists():
            await event.reply(f"❌ File yang Anda tentukan `{specified_filename}` tidak ditemukan.")
            return
        args.remove(file_arg)

    # Parsing baru untuk argumen limit opsional
    max_users_per_session = None
    limit_arg = next((arg for arg in args if arg.lower().startswith('limit=')), None)
    if limit_arg:
        try:
            max_users_per_session = int(limit_arg.split('=')[1])
            args.remove(limit_arg) # Hapus dari daftar argumen agar tidak mengganggu parsing lama
        except (ValueError, IndexError):
            await event.reply("❌ Format limit salah. Gunakan `limit=<angka>`, contoh: `limit=10`.")
            return
    
    is_daily_task = 'daily' in [arg.lower() for arg in args]
    if is_daily_task:
        args = [arg for arg in args if arg.lower() != 'daily']

    if len(args) < 3:
        await event.reply(f"❌ **Format Salah!**\n\nGunakan: `/addgrup <sesi> <target> <jeda> [link] [limit=<angka>]`")
        return
    
    # Cek apakah argumen terakhir adalah link
    if len(args) >= 4 and (args[-1].startswith('http://') or args[-1].startswith('https://')):
        invite_link = args[-1]
        delay_minutes_str = args[-2]
        target_str = args[-3]
        session_names_str = " ".join(args[:-3])
    else:
        invite_link = None
        delay_minutes_str = args[-1]
        target_str = args[-2]
        session_names_str = " ".join(args[:-2])

    try:
        delay_minutes = int(delay_minutes_str)
    except ValueError:
        await event.reply("❌ `<jeda_menit>` harus berupa angka.")
        return

    session_names = [s.strip() for s in session_names_str.split(',') if s.strip()]
    if not session_names:
        await event.reply("❌ **Format Salah!**\nNama sesi tidak boleh kosong.")
        return

    asyncio.create_task(run_pooled_broadcast_task(event, session_names, target_str, delay_minutes, invite_link, mode='default', excel_file_path=excel_file_path, start_message_id=event.message.id, max_users_per_session=max_users_per_session, daily=is_daily_task, history_filename=history_filename, retry_failed_only=retry_failed_only))

@bot_client.on(events.NewMessage(pattern=r'/addgrupfast (.*)'))
async def addgrupfast_handler(event):
    print(f"[INFO] Perintah '{event.raw_text}' dari user {event.sender_id} di chat {event.chat_id}")
    args = [arg for arg in event.pattern_match.group(1).split(' ') if arg]

    # Parsing untuk riwayat
    history_filename = None
    history_arg = next((arg for arg in args if arg.lower().startswith('history=')), None)
    if history_arg:
        history_filename = history_arg.split('=', 1)[1]
        args.remove(history_arg)

    # Parsing untuk retry
    retry_failed_only = 'retry=failed' in [arg.lower() for arg in args]
    if retry_failed_only:
        args = [arg for arg in args if arg.lower() != 'retry=failed']

    # Parsing untuk file spesifik
    excel_file_path = None
    file_arg = next((arg for arg in args if arg.lower().startswith('file=')), None)
    if file_arg:
        specified_filename = file_arg.split('=', 1)[1]
        excel_file_path = SCRIPT_DIR / specified_filename
        if not excel_file_path.exists():
            await event.reply(f"❌ File yang Anda tentukan `{specified_filename}` tidak ditemukan.")
            return
        args.remove(file_arg)

    # Parsing baru untuk argumen limit opsional
    max_users_per_session = None
    limit_arg = next((arg for arg in args if arg.lower().startswith('limit=')), None)
    if limit_arg:
        try:
            max_users_per_session = int(limit_arg.split('=')[1])
            args.remove(limit_arg) # Hapus dari daftar argumen
        except (ValueError, IndexError):
            await event.reply("❌ Format limit salah. Gunakan `limit=<angka>`, contoh: `limit=10`.")
            return
            
    is_daily_task = 'daily' in [arg.lower() for arg in args]
    if is_daily_task:
        args = [arg for arg in args if arg.lower() != 'daily']

    if len(args) < 3:
        await event.reply(f"❌ **Format Salah!**\n\nGunakan: `/addgrupfast <sesi> <target> <jeda> [limit=<angka>]`\n\nLihat /help untuk detail.")
        return

    try:
        delay_minutes = int(args[-1])
    except ValueError:
        await event.reply("❌ **Format Salah!**\n`<jeda_menit>` harus berupa angka.")
        return

    target_str = args[-2]
    session_names_str = " ".join(args[:-2])
    session_names = [s.strip() for s in session_names_str.split(',') if s.strip()]

    if not session_names:
        await event.reply("❌ **Format Salah!**\nNama sesi tidak boleh kosong.")
        return

    asyncio.create_task(run_pooled_broadcast_task(event, session_names, target_str, delay_minutes, invite_link=None, mode='fast', excel_file_path=excel_file_path, start_message_id=event.message.id, max_users_per_session=max_users_per_session, daily=is_daily_task, history_filename=history_filename, retry_failed_only=retry_failed_only))

@bot_client.on(events.NewMessage(pattern=r'/addgrupexcel (.*)'))
async def addgrupexcel_handler(event):
    print(f"[INFO] Perintah '{event.raw_text}' dari user {event.sender_id} di chat {event.chat_id}")
    args = [arg for arg in event.pattern_match.group(1).split(' ') if arg]

    # Parsing untuk riwayat
    history_filename = None
    history_arg = next((arg for arg in args if arg.lower().startswith('history=')), None)
    if history_arg:
        history_filename = history_arg.split('=', 1)[1]
        args.remove(history_arg)

    # Parsing untuk retry
    retry_failed_only = 'retry=failed' in [arg.lower() for arg in args]
    if retry_failed_only:
        args = [arg for arg in args if arg.lower() != 'retry=failed']

    # Parsing baru untuk argumen limit opsional
    max_users_per_session = None
    limit_arg = next((arg for arg in args if arg.lower().startswith('limit=')), None)
    if limit_arg:
        try:
            max_users_per_session = int(limit_arg.split('=')[1])
            args.remove(limit_arg) # Hapus dari daftar argumen
        except (ValueError, IndexError):
            await event.reply("❌ Format limit salah. Gunakan `limit=<angka>`, contoh: `limit=10`.")
            return
            
    is_daily_task = 'daily' in [arg.lower() for arg in args]
    if is_daily_task:
        args = [arg for arg in args if arg.lower() != 'daily']

    if len(args) < 3:
        await event.reply(f"❌ **Format Salah!**\n\nGunakan: `/addgrupexcel <sesi> <target> <jeda> [link] [limit=<angka>]`")
        return
    
    # Cek apakah argumen terakhir adalah link
    if len(args) >= 4 and (args[-1].startswith('http://') or args[-1].startswith('https://')):
        invite_link = args[-1]
        delay_minutes_str = args[-2]
        target_str = args[-3]
        session_names_str = " ".join(args[:-3])
    else:
        invite_link = None
        delay_minutes_str = args[-1]
        target_str = args[-2]
        session_names_str = " ".join(args[:-2])

    try:
        delay_minutes = int(delay_minutes_str)
    except ValueError:
        await event.reply("❌ `<jeda_menit>` harus berupa angka.")
        return

    session_names = [s.strip() for s in session_names_str.split(',') if s.strip()]
    if not session_names:
        await event.reply("❌ **Format Salah!**\nNama sesi tidak boleh kosong.")
        return

    try:
        async with bot_client.conversation(event.chat_id, timeout=300) as conv:
            await conv.send_message("📂 **Silakan unggah file Excel (.xlsx) Anda sekarang.**\n\nPastikan file memiliki kolom `uid` (User ID). Proses akan dibatalkan jika file tidak diunggah dalam 5 menit.")
            response = await conv.get_response()

            file_name = next((attr.file_name for attr in response.document.attributes if isinstance(attr, DocumentAttributeFilename)), None)

            if not response.document or not (response.document.mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' or (file_name and file_name.lower().endswith('.xlsx'))):
                await conv.send_message("❌ Unggahan tidak valid atau bukan file Excel (.xlsx). Proses dibatalkan.")
                return
            
            # Simpan file yang diunggah dengan nama sesi untuk konsistensi
            # Gunakan nama sesi pertama untuk nama file
            first_session = session_names[0] if session_names else "pool"
            download_path = SCRIPT_DIR / f"manual_upload_{first_session}_{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
            await conv.send_message(f"⏳ Mengunduh file `{file_name or 'file.xlsx'}`...")
            await bot_client.download_media(response.media, file=download_path)
            await conv.send_message("✅ File berhasil diunduh. Memulai proses penambahan anggota...")

            asyncio.create_task(run_pooled_broadcast_task(
                event, session_names, target_str, delay_minutes, 
                invite_link, mode='default', excel_file_path=download_path, 
                start_message_id=event.message.id, max_users_per_session=max_users_per_session, daily=is_daily_task,
                history_filename=history_filename, retry_failed_only=retry_failed_only
            ))
    except asyncio.TimeoutError:
        await event.reply("⏱️ Waktu tunggu untuk unggah file habis. Proses dibatalkan.")
    except Exception as e:
        await event.reply(f"❌ Terjadi error saat proses unggah file: {e}")
        traceback.print_exc()

@bot_client.on(events.NewMessage(pattern='/idgrup'))
async def idgrup_handler(event):
    print(f"[INFO] Perintah '{event.raw_text}' dari user {event.sender_id} di chat {event.chat_id}")
    # Periksa apakah perintah dijalankan di dalam grup atau channel
    if not event.is_group and not event.is_channel:
        await event.reply("❌ Perintah ini hanya bisa digunakan di dalam grup.")
        return

    try:
        # Dapatkan informasi chat saat ini
        chat = await event.get_chat()
        chat_id = event.chat_id
        chat_title = chat.title

        message = (
            f"📄 **Informasi Grup Ini:**\n\n"
            f"**Nama Grup:** {chat_title}\n"
            f"**ID Grup:** `{chat_id}`"
        )
        await event.reply(message, parse_mode='md')

    except Exception as e:
        await event.reply(f"❌ Terjadi error saat mengambil ID grup:\n`{e}`")
        traceback.print_exc()

@bot_client.on(events.NewMessage(pattern='/status'))
async def status_handler(event):
    print(f"[INFO] Perintah '{event.raw_text}' dari user {event.sender_id} di chat {event.chat_id}")
    running_tasks = {session: data for session, data in TASK_STATE.items() if data.get("running")}
    if not running_tasks:
        await event.reply("⏹️ Tidak ada proses yang sedang berjalan saat ini.")
        return
    message = "⚙️ **Status Proses yang Sedang Berjalan:**\n\n"
    for session, data in running_tasks.items():
        task_name = data.get('task_name', 'Tidak diketahui')
        message += f"🔹 Akun: `{session}` | Tugas: `{task_name}`\n"
    await event.reply(message)

@bot_client.on(events.NewMessage(pattern=r'/stop(?: (\w+))?'))
async def stop_handler(event):
    print(f"[INFO] Perintah '{event.raw_text}' dari user {event.sender_id} di chat {event.chat_id}")
    # Check if user is admin
    if event.sender_id != ADMIN_ID:
        await event.reply("❌ Anda tidak memiliki izin untuk menggunakan perintah ini.")
        return

    session_name = event.pattern_match.group(1)

    if session_name:
        # Hentikan tugas spesifik
        if not TASK_STATE.get(session_name, {}).get("running"):
            await event.reply(f"⏹️ Tidak ada proses yang berjalan untuk akun `{session_name}`.")
            return

        TASK_STATE[session_name]["stop_requested"] = True
        await event.reply(f"⏳ Perintah stop telah dikirim untuk tugas di akun `{session_name}`. Proses akan berhenti pada iterasi berikutnya.")
    else:
        # Hentikan semua tugas
        running_tasks = [s for s, data in TASK_STATE.items() if data.get("running")]
        if not running_tasks:
            await event.reply("⏹️ Tidak ada proses yang sedang berjalan saat ini.")
            return

        for s_name in running_tasks:
            if TASK_STATE.get(s_name):
                TASK_STATE[s_name]["stop_requested"] = True
        
        await event.reply(f"⏳ Perintah stop telah dikirim untuk semua ({len(running_tasks)}) tugas yang berjalan. Proses akan berhenti pada iterasi berikutnya.")

@bot_client.on(events.NewMessage(pattern='/leavegroup'))
async def leavegroup_handler(event):
    print(f"[INFO] Perintah '{event.raw_text}' dari user {event.sender_id} di chat {event.chat_id}")
    # Periksa apakah pengguna adalah admin
    if event.sender_id != ADMIN_ID:
        await event.reply("❌ Anda tidak memiliki izin untuk menggunakan perintah ini.")
        return

    # Periksa apakah perintah dijalankan di dalam grup
    if not event.is_group and not event.is_channel:
        await event.reply("❌ Perintah ini hanya bisa digunakan di dalam grup.")
        return

    try:
        chat = await event.get_chat()
        await event.reply(f"✅ Oke, saya akan keluar dari grup **{chat.title}**...")
        await asyncio.sleep(2)  # Jeda agar pesan terbaca
        await bot_client.delete_dialog(event.chat_id)
        print(f"[INFO] Bot berhasil keluar dari grup '{chat.title}' (ID: {event.chat_id}) atas perintah admin.")
    except Exception as e:
        await event.reply(f"❌ Terjadi error saat mencoba keluar dari grup:\n`{e}`")
        traceback.print_exc()

# =====================================================
# TITIK MASUK UTAMA
# =====================================================

async def main():
    """Fungsi utama untuk menjalankan bot."""
    # Pemeriksaan ini akan mencegah warning pada proses start berikutnya.
    # Metode start() akan menghubungkan klien dan login jika perlu.
    # Ini akan menggunakan sesi yang ada jika sudah valid, yang mungkin menampilkan peringatan yang tidak berbahaya.
    await bot_client.start(bot_token=BOT_TOKEN)
    print("✅ Bot berhasil online dan siap menerima perintah.")
    print(f"🔐 Pastikan ADMIN_ID diatur dengan benar untuk menggunakan perintah admin.")
    await bot_client.run_until_disconnected()

if __name__ == "__main__":
    print("╔════════════════════════════════════════════════════╗")
    print("║   TELEGRAM GROUP SCRAPER & BROADCASTER BOT (MULTI-AKUN)   ║")
    print("╚════════════════════════════════════════════════════╝")

    if not all([API_ID, API_HASH, BOT_TOKEN, ADMIN_ID]):
        print("\n\n⚠️  PERINGATAN: Satu atau lebih variabel (API_ID, API_HASH, BOT_TOKEN, ADMIN_ID) tidak diatur dalam file .env Anda.")
        print("     Pastikan file .env ada dan berisi semua nilai yang diperlukan.")
        print("     Bot akan tetap berjalan, tetapi Anda tidak akan bisa menggunakan perintah admin.\n")

    # Jalankan loop utama bot
    try:
        # main() akan menangani seluruh siklus hidup koneksi bot.
        bot_client.loop.run_until_complete(main())
    except KeyboardInterrupt:
        print("\n⏹️ Bot dihentikan.")
    except Exception as e:
        print(f"\n❌ Error fatal: {e}")
        traceback.print_exc()
